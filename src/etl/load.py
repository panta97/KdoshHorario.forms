import os
import shutil
import openpyxl
import xlwings as xw
import pandas as pd
import numpy as np

from win32com.client import Dispatch
from src.general.base import XlsRange


def create_missing_sheets(template_filepath, new_filepath, sheetnames):
    xl = Dispatch('Excel.Application')

    template_wb = xl.Workbooks.Open(Filename=template_filepath)
    new_wb = xl.Workbooks.Open(Filename=new_filepath)
    template_ws = template_wb.Worksheets(1)

    for _ in range(1, len(sheetnames)):
        template_ws.Copy(Before=new_wb.Worksheets(1))

    # RENAME SHEETNAMES
    for i, sheetname in enumerate(sheetnames, start=1):
        new_wb.Worksheets(i).Name = sheetname

    template_wb.Close(SaveChanges=False)
    new_wb.Close(SaveChanges=True)

def create_new_xlsm(filename, sheetnames):
    dir_path = os.path.dirname(os.path.realpath(__file__))
    dir_path = dir_path[:-4] # TODO: HARDCODED
    filepath = os.path.join(dir_path, '_BASE.xlsm')
    shutil.copy(filepath, filename)

    create_missing_sheets(filepath, filename, sheetnames)

def generate_schedule_templates(filepath, filename, dfs, areas, sheetnames):
    # AREA[0] ID
    # AREA[1] NAME

    filepath = os.path.join(filepath, filename)
    create_new_xlsm(filepath, sheetnames) # TODO: SHEETNAMES ARE HARDCODED

    app = xw.App(visible=False)
    wb_generic = app.books.open(filepath)

    for i, area in enumerate(areas):
        sheetname = area[1]
        work_sheet = wb_generic.sheets[sheetname]
        
        rm = XlsRange(1, 1, 149, 38)

        val_df = dfs[i].values
        work_sheet.range((rm.frow, rm.fcol), (rm.lrow, rm.lcol)).value = val_df

    wb_generic.save()
    wb_generic.close()
    app.quit()